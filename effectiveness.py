import numpy as np
import openpyxl
from openpyxl.styles import Alignment
import os

class ArtificialBeeColony:
    def __init__(self, objective_func, dim_p, dim_y, constraints, pop_size, max_iter, M, N, alpha, beta, gamma, x, w):
        self.objective_func = objective_func
        self.dim_p = dim_p
        self.dim_y = dim_y
        self.constraints = constraints
        self.pop_size = pop_size
        self.max_iter = max_iter
        self.M = M
        self.N = N
        self.alpha = alpha
        self.beta = beta
        self.gamma = gamma
        self.x = x
        self.w = w

    def optimize(self):
        population_p = self.init_population_p()
        population_y = self.init_population_y()
        best_solution_p, best_solution_y, best_fitness = self.run_optimization(population_p, population_y)
        return best_solution_p, best_solution_y, best_fitness

    def local_search(self, solution_p, solution_y):
        best_p, best_y = solution_p.copy(), solution_y.copy()
        best_fitness = self.objective_func(best_p, best_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w)

        for _ in range(1000):
            new_p, new_y = self.generate_new_solution(best_p, best_y)
            if self.is_feasible(new_p, new_y):
                new_fitness = self.objective_func(new_p, new_y, self.M, self.N, self.alpha, self.beta, self.gamma,
                                                  self.w)
                if new_fitness > best_fitness:
                    best_p, best_y = new_p, new_y
                    best_fitness = new_fitness
                else:
                    break

        return best_p, best_y, best_fitness

    def run_optimization(self, population_p, population_y):
        best_solution_p = None
        best_solution_y = None
        best_fitness = float('-inf')
        scout_bee_ratio = 0.25

        for i in range(self.max_iter):
            print(f"Iteration {i + 1}:")
            population_p, population_y = self.perform_employed_bee_phase(population_p, population_y)
            probabilities = self.calculate_probabilities(population_p, population_y)
            population_p, population_y = self.perform_onlooker_bee_phase(population_p, population_y, probabilities)
            population_p, population_y = self.perform_scout_bee_phase(population_p, population_y, scout_bee_ratio)


            for j in range(self.pop_size):
                population_p[j], population_y[j], _ = self.local_search(population_p[j], population_y[j])

            best_solution_p, best_solution_y, best_fitness = self.update_best_solution(population_p, population_y,
                                                                                       best_solution_p, best_solution_y,
                                                                                       best_fitness)
            print(f"Best fitness so far: {best_fitness}")

        return best_solution_p, best_solution_y, best_fitness

    def perform_employed_bee_phase(self, population_p, population_y):
        for j in range(self.pop_size):
            new_solution_p, _ = self.generate_new_solution(population_p[j], population_y[j])
            if self.is_feasible(new_solution_p, population_y[j]):
                new_fitness = self.objective_func(new_solution_p, population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                if new_fitness > self.objective_func(population_p[j], population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w):
                    population_p[j] = new_solution_p
        return population_p, population_y

    def perform_onlooker_bee_phase(self, population_p, population_y, probabilities):
        for j in range(self.pop_size):
            selected_solution_p, selected_solution_y = self.select_solution(population_p, population_y, probabilities)
            new_solution_p, _ = self.generate_new_solution(selected_solution_p, selected_solution_y)
            if self.is_feasible(new_solution_p, selected_solution_y):
                new_fitness = self.objective_func(new_solution_p, selected_solution_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                if new_fitness > self.objective_func(selected_solution_p, selected_solution_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w):
                    selected_solution_p[:] = new_solution_p
        return population_p, population_y

    def perform_scout_bee_phase(self, population_p, population_y, scout_bee_ratio):
        for j in range(self.pop_size):
            if np.random.rand() < scout_bee_ratio:
                new_solution_p, new_solution_y = self.generate_new_solution(population_p[j], population_y[j])
                if self.is_feasible(new_solution_p, new_solution_y):
                    new_fitness_p = self.objective_func(new_solution_p, population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                    new_fitness_y = self.objective_func(population_p[j], new_solution_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                    if new_fitness_p > self.objective_func(population_p[j], population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w):
                        population_p[j] = new_solution_p
                    if new_fitness_y > self.objective_func(population_p[j], population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w):
                        population_y[j] = new_solution_y
        return population_p, population_y

    def update_best_solution(self, population_p, population_y, best_solution_p, best_solution_y, best_fitness):
        for j in range(self.pop_size):
            if self.is_feasible(population_p[j], population_y[j]):
                fitness = self.objective_func(population_p[j], population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                if fitness > best_fitness:
                    best_solution_p = population_p[j]
                    best_solution_y = population_y[j]
                    best_fitness = fitness
        return best_solution_p, best_solution_y, best_fitness

    def init_population_p(self):
        population = np.zeros((self.pop_size, self.dim_p))
        for i in range(self.pop_size):
            for j in range(self.dim_p):
                population[i, j] = np.random.uniform(self.constraints['p_min'][j // self.N],
                                                     self.constraints['p_max'][j // self.N])
                if np.random.rand() < 0.2:  # 20% 概率重新初始化
                    population[i, j] = np.random.uniform(self.constraints['p_min'][j // self.N],
                                                         self.constraints['p_max'][j // self.N])
        return population

    def init_population_y(self):
        population = []
        for _ in range(self.pop_size):
            y = np.zeros((self.M, self.N), dtype=int)
            rows = np.random.choice(self.M, size=self.N, replace=False)
            cols = np.random.choice(self.N, size=self.M, replace=False)
            y[np.arange(self.M), cols] = 1
            y = y.flatten()
            population.append(y)
        return np.array(population)

    def generate_new_solution(self, solution_p, solution_y):
        new_solution_p = solution_p.copy()
        new_solution_y = self.generate_random_feasible_y()
        dim_p = np.random.randint(self.dim_p)
        new_solution_p[dim_p] = np.random.uniform(self.constraints['p_min'][dim_p // self.N],
                                                  self.constraints['p_max'][dim_p // self.N])
        return new_solution_p, new_solution_y

    def generate_random_solution(self):
        solution_p = np.zeros(self.dim_p)
        for i in range(self.dim_p):
            solution_p[i] = np.random.uniform(self.constraints['p_min'][i // self.N], self.constraints['p_max'][i // self.N])
        solution_y = np.random.randint(0, 2, size=self.dim_y)
        return solution_p, solution_y
    def generate_random_feasible_y(self):
        y = np.zeros((self.M, self.N), dtype=int)
        rows = np.random.choice(self.M, size=self.N, replace=False)
        cols = np.random.choice(self.N, size=self.M, replace=False)
        y[np.arange(self.M), cols] = 1
        return y.flatten()

    def calculate_probabilities(self, population_p, population_y):
        fitness_values = np.zeros(self.pop_size)
        for i in range(self.pop_size):
            if self.is_feasible(population_p[i], population_y[i]):
                fitness_values[i] = self.objective_func(population_p[i], population_y[i], self.M, self.N,
                                                            self.alpha, self.beta, self.gamma, self.w)
        probabilities = fitness_values / np.sum(fitness_values)
        return probabilities

    def select_solution(self, population_p, population_y, probabilities):
        index = np.random.choice(range(self.pop_size), p=probabilities)
        return population_p[index], population_y[index]

    def is_feasible(self, solution_p, solution_y):
        y = solution_y.reshape((self.M, self.N))
        p = solution_p.reshape((self.M, self.N))

        # Constraint 1: sum(y_ij) <= 1 for each i
        if np.any(np.sum(y, axis=1) > 1):
            return False

        # Constraint 2: sum(y_ij) <= 1 for each j
        if np.any(np.sum(y, axis=0) > 1):
            return False

        # Constraint 3: y_ij in {0, 1}
        if np.any((y != 0) & (y != 1)):
            return False

        # Constraint 4: p_ij <= p_i^u
        for i in range(self.M):
            if np.any(p[i, :] > self.constraints['p_max'][i]):
                return False

        # Constraint 5: p_ij >= p_i^l
        for i in range(self.M):
            if np.any(p[i, :] < self.constraints['p_min'][i]):
                return False

        return True

    def print_selected_values(self, y, p, w):
        print("Selected values:")
        for i in range(self.M):
            for j in range(self.N):
                if y[i, j] == 1:
                    print(f"p[{i},{j}] = {p[i, j]}, w[{i},{j}] = {w[i, j]}")

    def apply_laplacian_noise(self, values):
        noise = np.random.laplace(loc=0.0, scale=1/self.M, size=len(values))
        return values + noise

    def apply_bounded_laplacian_noise(self, values, min_values, max_values):
        noise = np.random.laplace(loc=0.0, scale=1/self.M, size=len(values))
        noisy_values = values + noise
        noisy_values = np.clip(noisy_values, min_values, max_values)
        return noisy_values

    def calculate_new_objective(self, best_p, best_y, w):
        noisy_p_unbounded = self.apply_laplacian_noise(best_p.flatten())
        noisy_p_unbounded = noisy_p_unbounded.reshape((self.M, self.N))
        noisy_w_unbounded = self.apply_laplacian_noise(w.flatten())
        noisy_w_unbounded = noisy_w_unbounded.reshape((self.M, self.N))

        new_objective_unbounded = 0
        for i in range(self.M):
            exponent = 0
            for j in range(self.N):
                exponent += (self.alpha[i] - self.beta[i] * noisy_p_unbounded[i, j] - self.gamma[i] * noisy_w_unbounded[
                    i, j]) * best_y[i, j]
            new_objective_unbounded += np.sum(
                noisy_p_unbounded[i, :] * best_y[i, :] * np.exp(exponent) / (1 + np.exp(exponent)))

        p_min_reshaped = np.repeat(self.constraints['p_min'], self.N)
        p_max_reshaped = np.repeat(self.constraints['p_max'], self.N)
        noisy_p_bounded = self.apply_bounded_laplacian_noise(best_p.flatten(), p_min_reshaped, p_max_reshaped)
        noisy_p_bounded = noisy_p_bounded.reshape((self.M, self.N))
        noisy_w_bounded = self.apply_bounded_laplacian_noise(w.flatten(), 0, 5)
        noisy_w_bounded = noisy_w_bounded.reshape((self.M, self.N))

        new_objective_bounded = 0
        for i in range(self.M):
            exponent = 0
            for j in range(self.N):
                exponent += (self.alpha[i] - self.beta[i] * noisy_p_bounded[i, j] - self.gamma[i] * noisy_w_bounded[
                    i, j]) * best_y[i, j]
            new_objective_bounded += np.sum(
                noisy_p_bounded[i, :] * best_y[i, :] * np.exp(exponent) / (1 + np.exp(exponent)))

        return new_objective_unbounded, new_objective_bounded

    def get_final_objective(self, best_solution_y, best_solution_p):
        best_y = best_solution_y.reshape((self.M, self.N))

        final_objective = 0
        for i in range(self.M):
            exponent = 0
            for j in range(self.N):
                p_ij = self.constraints['p_min'][i]
                exponent += (self.alpha[i] - self.beta[i] * p_ij - self.gamma[i] * self.w[i, j]) * best_y[i, j]
            final_objective += np.sum(p_ij * best_y[i, :] * np.exp(exponent) / (1 + np.exp(exponent)))
        return final_objective

def objective(solution_p, solution_y, M, N, alpha, beta, gamma, w):
    y = solution_y.reshape((M, N))
    p = solution_p.reshape((M, N))

    result = 0
    for i in range(M):
        exponent = 0
        for j in range(N):
            exponent += (alpha[i] - beta[i] * p[i, j] - gamma[i] * w[i, j]) * y[i, j]
        result += np.sum(p[i, :] * y[i, :] * np.exp(exponent) / (1 + np.exp(exponent)))

    return result



def main():

    file_name = " "
    workbook = openpyxl.load_workbook(file_name)

    df1 = pd.read_excel(file_path, sheet_name="Sheet1")  # 读取 Sheet1
    df2 = pd.read_excel(file_path, sheet_name="Sheet2")  # 读取 Sheet2
    M = len(df1)
    N = len(df2)


    results_workbook = openpyxl.Workbook()
    results_sheet = results_workbook.active
    results_sheet.title = "Experiment Results-70"


    results_sheet["A1"] = "Experiment"
    results_sheet["B1"] = "Best Fitness"
    results_sheet["C1"] = "Final Objective"
    results_sheet["D1"] = "Unbounded Laplacian Noise Difference"
    results_sheet["E1"] = "Bounded Laplacian Noise Difference"
    results_sheet.row_dimensions[1].height = 20
    results_sheet.column_dimensions['A'].width = 15
    results_sheet.column_dimensions['B'].width = 15
    results_sheet.column_dimensions['C'].width = 15
    results_sheet.column_dimensions['D'].width = 25
    results_sheet.column_dimensions['E'].width = 25
    results_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    results_sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")
    results_sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")
    results_sheet["D1"].alignment = Alignment(horizontal="center", vertical="center")
    results_sheet["E1"].alignment = Alignment(horizontal="center", vertical="center")


    for i in range(5):

        sheet1 = workbook["Sheet1"]
        alpha = []
        beta = []
        gamma = []
        p_min = []
        p_max = []

        for row in sheet1.iter_rows(min_row=1, min_col=1, max_col=5, values_only=True):
            p_min.append(row[3])
            p_max.append(row[4])
            alpha.append(np.random.randint(int(p_min[-1] * 1.5), int(p_min[-1] * 1.5 + 30)))
            beta.append(np.round(np.random.uniform(0.8, 2.0), 1))
            gamma.append(np.random.randint(1, 9))

        alpha = np.array(alpha)
        beta = np.array(beta)
        gamma = np.array(gamma)
        p_min = np.array(p_min)
        p_max = np.array(p_max)

        sheet2 = workbook["Sheet2"]
        w = []
        for row in sheet2.iter_rows(min_row=1, min_col=1, max_col=70, values_only=True):
            w.append(list(np.random.randint(1, 8, size=70)))
        w = np.array(w)

        x = np.ones(M)

        dim_p = M * N
        dim_y = M * N

        constraints = {
            'p_min': p_min,
            'p_max': p_max
        }

        pop_size = 150

        max_iter = 2000

        abc = ArtificialBeeColony(objective, dim_p, dim_y, constraints, pop_size, max_iter, M, N, alpha, beta, gamma, x, w)

        best_solution_p, best_solution_y, best_fitness = abc.optimize()

        best_p = best_solution_p.reshape((M, N))
        best_y = best_solution_y.reshape((M, N))

        original_objective = abc.objective_func(best_p, best_y, M, N, alpha, beta, gamma, w)
        new_objective_unbounded, new_objective_bounded = abc.calculate_new_objective(best_p, best_y, w)
        objective_difference_unbounded = new_objective_unbounded - original_objective
        objective_difference_bounded = new_objective_bounded - original_objective

        final_objective = abc.get_final_objective(best_y, best_p)

        results_sheet.cell(row=i+2, column=1, value=i+1)
        results_sheet.cell(row=i+2, column=2, value=best_fitness)
        results_sheet.cell(row=i+2, column=3, value=final_objective)
        results_sheet.cell(row=i+2, column=4, value=objective_difference_unbounded)
        results_sheet.cell(row=i+2, column=5, value=objective_difference_bounded)

    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "experiment_results-peak.xlsx")
    results_workbook.save(desktop_path)
    print(f"Results saved to: {desktop_path}")

if __name__ == "__main__":
    main()

