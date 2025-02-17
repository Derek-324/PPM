import numpy as np
import openpyxl

class ArtificialBeeColony:
    def __init__(self, objective_func, dim_p, dim_y, constraints, pop_size, max_iter, M, N, alpha, beta, gamma, x, w):
        self.objective_func = objective_func
        self.dim_p = dim_p
        self.dim_y = dim_y
        self.constraints = constraints
        self.pop_size = pop_size
        self.max_iter = max_iter
        self.M = M
        self.N = N
        self.alpha = alpha
        self.beta = beta
        self.gamma = gamma
        self.x = x
        self.w = w
        self.p_min = constraints['p_min']
        self.p_max = constraints['p_max']

    def optimize(self):
        population_p = self.init_population_p()
        population_y = self.init_population_y()
        best_solution_p, best_solution_y, best_fitness = self.run_optimization(population_p, population_y)
        return best_solution_p, best_solution_y, best_fitness

    def run_optimization(self, population_p, population_y):
        best_solution_p = None
        best_solution_y = None
        best_fitness = float('-inf')
        scout_bee_ratio = 0.3

        for i in range(self.max_iter):
            print(f"Iteration {i + 1}:")
            population_p, population_y = self.perform_employed_bee_phase(population_p, population_y)
            probabilities = self.calculate_probabilities(population_p, population_y)
            population_p, population_y = self.perform_onlooker_bee_phase(population_p, population_y, probabilities)
            population_p, population_y = self.perform_scout_bee_phase(population_p, population_y, scout_bee_ratio)

            best_solution_p, best_solution_y, best_fitness = self.update_best_solution(population_p, population_y,
                                                                                       best_solution_p, best_solution_y,
                                                                                       best_fitness)
            print(f"Best fitness so far: {best_fitness}")

        return best_solution_p, best_solution_y, best_fitness

    def perform_employed_bee_phase(self, population_p, population_y):
        for j in range(self.pop_size):
            new_solution_p, new_solution_y = self.generate_new_solution(population_p[j], population_y[j])
            if self.is_feasible(new_solution_p, new_solution_y):
                new_fitness = self.objective_func(new_solution_p, new_solution_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                if new_fitness > self.objective_func(population_p[j], population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w):
                    population_p[j] = new_solution_p
                    population_y[j] = new_solution_y
        return population_p, population_y

    def perform_onlooker_bee_phase(self, population_p, population_y, probabilities):
        for j in range(self.pop_size):
            selected_solution_p, selected_solution_y = self.select_solution(population_p, population_y, probabilities)
            new_solution_p, new_solution_y = self.generate_new_solution(selected_solution_p, selected_solution_y)
            if self.is_feasible(new_solution_p, new_solution_y):
                new_fitness = self.objective_func(new_solution_p, new_solution_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                if new_fitness > self.objective_func(selected_solution_p, selected_solution_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w):
                    selected_solution_p[:] = new_solution_p
                    selected_solution_y[:] = new_solution_y
        return population_p, population_y

    def perform_scout_bee_phase(self, population_p, population_y, scout_bee_ratio):
        for j in range(self.pop_size):
            if np.random.rand() < scout_bee_ratio:
                new_solution_p, new_solution_y = self.generate_random_solution()
                if self.is_feasible(new_solution_p, new_solution_y):
                    new_fitness = self.objective_func(new_solution_p, new_solution_y, self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                    if new_fitness > self.objective_func(population_p[j], population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w):
                        population_p[j] = new_solution_p
                        population_y[j] = new_solution_y
        return population_p, population_y

    def update_best_solution(self, population_p, population_y, best_solution_p, best_solution_y, best_fitness):
        for j in range(self.pop_size):
            if self.is_feasible(population_p[j], population_y[j]):
                fitness = self.objective_func(population_p[j], population_y[j], self.M, self.N, self.alpha, self.beta, self.gamma, self.w)
                if fitness > best_fitness:
                    best_solution_p = population_p[j]
                    best_solution_y = population_y[j]
                    best_fitness = fitness
        return best_solution_p, best_solution_y, best_fitness

    def apply_bounded_laplacian_noise(self, values, min_values, max_values, laplacian_scale):
        loc = 0.0
        noise = np.random.laplace(loc=loc, scale=laplacian_scale, size=values.shape)
        noisy_values = values + noise
        noisy_values = np.clip(noisy_values, min_values, max_values)
        return noisy_values

    def init_population_p(self):
        population = np.zeros((self.pop_size, self.dim_p))
        for i in range(self.pop_size):
            for j in range(self.dim_p):
                population[i, j] = np.random.uniform(self.p_min[j // self.N],
                                                     self.p_max[j // self.N])
                if np.random.rand() < 0.2:  # 20% 概率重新初始化
                    population[i, j] = np.random.uniform(self.p_min[j // self.N],
                                                         self.p_max[j // self.N])
        return population

    def init_population_y(self):
        population = []
        for _ in range(self.pop_size):
            y = np.zeros((self.M, self.N), dtype=int)
            rows = np.random.choice(self.M, size=self.N, replace=False)
            cols = np.random.choice(self.N, size=self.M, replace=False)
            y[np.arange(self.M), cols] = 1
            y = y.flatten()
            population.append(y)
        return np.array(population)

    def generate_new_solution(self, solution_p, solution_y):
        new_solution_p = solution_p.copy()
        new_solution_y = solution_y.copy()
        dim_p = np.random.randint(self.dim_p)
        new_solution_p[dim_p] = np.random.uniform(self.p_min[dim_p // self.N],
                                                  self.p_max[dim_p // self.N])
        dim_y = np.random.randint(self.dim_y)
        new_solution_y[dim_y] = 1 - new_solution_y[dim_y]
        return new_solution_p, new_solution_y

    def generate_random_solution(self):
        solution_p = np.zeros(self.dim_p)
        for i in range(self.dim_p):
            solution_p[i] = np.random.uniform(self.p_min[i // self.N],
                                              self.p_max[i // self.N])
        solution_y = np.random.randint(0, 2, size=self.dim_y)
        return solution_p, solution_y

    def calculate_probabilities(self, population_p, population_y):
        fitness_values = np.zeros(self.pop_size)
        for i in range(self.pop_size):
            if self.is_feasible(population_p[i], population_y[i]):
                fitness_values[i] = self.objective_func(population_p[i], population_y[i], self.M, self.N,
                                                        self.alpha, self.beta, self.gamma, self.w)
        probabilities = fitness_values / np.sum(fitness_values)
        return probabilities

    def select_solution(self, population_p, population_y, probabilities):
        index = np.random.choice(range(self.pop_size), p=probabilities)
        return population_p[index], population_y[index]

    def is_feasible(self, solution_p, solution_y):
        y = solution_y.reshape((self.M, self.N))
        p = solution_p.reshape((self.M, self.N))

        # Constraint 1: sum(y_ij) <= 1 for each i
        if np.any(np.sum(y, axis=1) > 1):
            return False

        # Constraint 2: sum(y_ij) <= 1 for each j
        if np.any(np.sum(y, axis=0) > 1):
            return False

        # Constraint 3: y_ij in {0, 1}
        if np.any((y != 0) & (y != 1)):
            return False

        # Constraint 4: p_ij <= p_i^u
        for i in range(self.M):
            if np.any(p[i, :] > self.p_max[i]):
                return False

        # Constraint 5: p_ij >= p_i^l
        for i in range(self.M):
            if np.any(p[i, :] < self.p_min[i]):
                return False

        return True

def objective(solution_p, solution_y, M, N, alpha, beta, gamma, w):
    y = solution_y.reshape((M, N))
    p = solution_p.reshape((M, N))

    result = 0
    for i in range(M):
        exponent = 0
        for j in range(N):
            exponent += (alpha[i] - beta[i] * p[i, j] - gamma[i] * w[i, j]) * y[i, j]
        result += np.sum(p[i, :] * y[i, :] * np.exp(exponent) / (1 + np.exp(exponent)))

    return result

def main():
    file_name = " "
    workbook = openpyxl.load_workbook(file_name)

    df1 = pd.read_excel(file_path, sheet_name="Sheet1")  # 读取 Sheet1
    df2 = pd.read_excel(file_path, sheet_name="Sheet2")  # 读取 Sheet2
    M = len(df1)
    N = len(df2)

    # Read from Sheet1
    sheet1 = workbook["Sheet1"]
    alpha = []
    beta = []
    gamma = []
    p_min = []

    for row in sheet1.iter_rows(min_row=1, min_col=1, max_col=4, values_only=True):
        alpha.append(row[0])
        beta.append(row[1])
        gamma.append(row[2])
        p_min.append(row[3])

    alpha = np.array(alpha)
    beta = np.array(beta)
    gamma = np.array(gamma)
    p_min = np.array(p_min)

    # Read from Sheet2
    sheet2 = workbook["Sheet2"]
    w = []
    for row in sheet2.iter_rows(min_row=1, min_col=1, max_col=40, values_only=True):
        w.append(list(row))
    w = np.array(w)

    x = np.ones(M)

    # Dimensions
    dim_p = M * N
    dim_y = M * N

    # Population size
    pop_size = 1000

    # Maximum iterations
    max_iter = 2000

    # Create a new Excel file to store the results
    output_file = openpyxl.Workbook()
    output_sheet = output_file.active

    # Iterate over different p_max values
    for p_max_factor in [1.1, 1.2, 1.3, 1.4, 1.5]:
        p_max = p_min * p_max_factor

        # Constraint conditions
        constraints = {'p_min': p_min, 'p_max': p_max}

        # Create an instance of the ArtificialBeeColony class
        abc = ArtificialBeeColony(objective, dim_p, dim_y, constraints, pop_size, max_iter, M, N, alpha, beta, gamma, x,
                                  w)

        # Optimize and get the best solution
        best_solution_p, best_solution_y, best_fitness = abc.optimize()

        # Calculate the noisy objective with the best solution
        laplacian_scale = np.repeat((p_max - p_min) / 4, N)
        min_values = np.repeat(p_min, N)
        max_values = np.repeat(p_max, N)
        noisy_p = abc.apply_bounded_laplacian_noise(best_solution_p.flatten(), min_values, max_values, laplacian_scale)
        noisy_p = noisy_p.reshape((M, N))
        noisy_objective = objective(noisy_p.flatten(), best_solution_y, M, N, alpha, beta, gamma, w)

        # Write the results to the Excel file
        output_sheet.append([p_max_factor, best_fitness, noisy_objective])

    # Save the Excel file
    output_file.save("output.xlsx")

    # Save the Excel file
    output_file.save("output.xlsx")

if __name__ == "__main__":
    main()
